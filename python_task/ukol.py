
## ANETA JIRKOVA - PYTHON TASK FOR THE JOB INTERVIEW ON THURSDAY 20. July 2023 ##
## COMPANY: DTSE


## PART 1: 
# Task is to build solution in python for downloading and presenting covid-19 cases data.
# Data is available on this URL: https://covid.ourworldindata.org/data/owid-covid-data.xlsx
# Download the data within the script.

#Use requests lib for downloading from web
import requests
#use openpyxl for openXML 
from openpyxl import load_workbook
#use pandas for data procesing
import pandas as pd

#for graphs
import matplotlib.pyplot as plt
import seaborn



#set proxy if any
proxies = {
  # 'http': 'http://proxy.example.com:8080',
}
url = 'https://covid.ourworldindata.org/data/owid-covid-data.xlsx'


#download data using requests, save to local file
response = requests.get(url, proxies=proxies)
open("owid-covid-data.xlsx", "wb").write(response.content)



# load open XML document
#wb = load_workbook(filename = 'owid-covid-data_1.xlsx')
wb = load_workbook(filename = 'owid-covid-data.xlsx')
# load first sheet
ws = wb[wb.sheetnames[0]]

data = pd.DataFrame(ws.values)


# Make the first data row as column header
data.columns = data.iloc[0]

# Drop the first row
data = data.iloc[1:]

# Reset the DataFrame index
data.reset_index(drop=True, inplace=True)

##
## PART 2:
## Calculate average new covid-19 cases in one week (last seven days) per 100 000 people for each state.
##

data['new_cases'] = pd.to_numeric(data['new_cases'], errors='coerce')

# Calculate the mean of 'new_cases' column
avg_new_cases = data['new_cases'].mean()

# Make sure the 'date' column is in datetime format
data['date'] = pd.to_datetime(data['date'])

week_ago = data['date'].max() - pd.DateOffset(days=7)
last_week_data = data[data['date'] > week_ago]

# Calculate the average new cases in the last week for each location
weekly_cases_per_locations = last_week_data.groupby('location')['new_cases'].sum()

#merge two DataFrames
population_per_locality = pd.merge(data[['location', 'population']].drop_duplicates(), weekly_cases_per_locations, on='location')

# Calculate average new covid-19 cases in one week (last seven days) per 100 000 people for each state
population_per_locality_per_100k =population_per_locality['new_cases'] * 100000 / population_per_locality['population'];

# Print the result
for i in range(0, len(population_per_locality['location'])):
   print(f"Average new cases in the last 7 days in {population_per_locality['location'].values[i] } per 100k : {population_per_locality_per_100k.values[i]} ")


##
## PART 3:
## Show two bar plots next to each other, showing data of 20 states with most average new cases in a week per 100k people. Left bar will show population of states, right bar will show cases per 100k people in week.
##


#Filtering 20 states with most average new cases in a week
population_per_locality['AvgNewCasesPer100k'] = population_per_locality['new_cases'] * 100000 / population_per_locality['population'];
df_top20 = population_per_locality.sort_values('AvgNewCasesPer100k', ascending=False).head(20)

#Defining the bar plots
fig,ax = plt.subplots(nrows=1,ncols=2,figsize=(14,6))

#Bar plot showing population of states
ax[0].bar(list(df_top20['location']), list(df_top20['population']), color ='maroon',width = 0.4)
ax[0].set_title('Population of Top 20 States')
ax[0].set_xlabel('State')
ax[0].set_ylabel('Population')
ax[0].tick_params(axis='x', rotation=90)
# "High income" among the states. Mistake in the data.

#Bar plot showing cases per 100k people in week
ax[1].bar(list(df_top20['location']), list(df_top20['new_cases']), color ='blue',width = 0.4)
ax[1].set_title('Avg New Cases per 100k in a Week for Top 20 States')
ax[1].set_xlabel('State')
ax[1].set_ylabel('Avg New Cases per 100k in a Week')
ax[1].tick_params(axis='x', rotation=90)

plt.tight_layout()
plt.show()

##
## PART 4
## Show scatter plot showing relation between population and average new cases in seven days.
# Here we are interested in 20 states with most average new cases in a week.

# Filter the data
weekly_cases_per_locations_avg = last_week_data.groupby('location')['new_cases'].mean()
population_per_locality['week_average'] = weekly_cases_per_locations_avg.values
avg_top20 = population_per_locality.sort_values('week_average', ascending=False).head(20)

# Scatter plot showing relation between population and average new cases in seven days
g = seaborn.JointGrid(data=avg_top20, x='population', y='week_average')
g.plot_joint(seaborn.scatterplot, legend=False)
plt.show()